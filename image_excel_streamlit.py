import requests
from tempfile import NamedTemporaryFile
import streamlit as st
from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import math
import io
from functools import lru_cache

# ========================
# 常量定义
# ========================
COLOR_DICT = {
    "#F9F9F2": 1,
    "#000000": 2,
    "#FFA44B": 3,
    "#FDA51B": 4,
    "#FFC8B7": 6,
    "#FD98BB": 7,
    "#F9DED4": 8,
    "#FFD8AD": 9,
    "#FA5C06": 10,
    "#F2A798": 11,
    "#E7C2B3": 12,
    "#F70F25": 13,
    "#CF638E": 14,
    "#F6545C": 15,
    "#FFBA77": 16,
    "#DE70D0": 17,
    "#F16C3D": 18,
    "#EE8A40": 19,
    "#FAB31A": 20,
    "#C8EF0F": 21,
    "#FE5300": 22,
    "#DF8561": 23,
    "#019EEF": 24,
    "#FA7A65": 25,
    "#F4986A": 26,
    "#F69E41": 27,
    "#F8F47A": 28,
    "#FFFF00": 29,
    "#CA6026": 30,
    "#C91C35": 31,
    "#97605A": 32,
    "#D2D1C4": 33,
    "#9A9898": 34,
    "#D090FD": 35,
    "#DEC8EF": 36,
    "#C59EE2": 37,
    "#C073FB": 38,
    "#D004B1": 39,
    "#5F2989": 41,
    "#FAE710": 42,
    "#E240CA": 43,
    "#EC3CC9": 44,
    "#FBC9FE": 45,
    "#F5B2FA": 46,
    "#FEE9DD": 47,
    "#BFF0EC": 48,
    "#6BDBCF": 49,
    "#A9DDEB": 50,
    "#FFFBDB": 51,
    "#82B8FA": 52,
    "#FEC389": 53,
    "#F0A17E": 54,
    "#2E87F4": 55,
    "#BFBFBF": 56,
    "#FE342A": 57,
    "#0234B4": 58,
    "#5A1D1C": 59,
    "#048696": 60,
    "#F4F579": 61,
    "#60E101": 62,
    "#3AE76E": 63,
    "#54ED11": 64,
    "#0560F7": 65,
    "#14C414": 66,
    "#4A4949": 69,
    "#FEB5D4": 70,
    "#00B050": 71,
    "#3862CC": 75,
    "#7D0024": 78,
    "#6B3618": 79,
    "#EE822F": 86,
    "#FFFFFF": 87,
    "#EEF2FB": 88,
    "#DEDEDE": 89,
    "#EED7C5": 92,
    "#D67752": 95,
    "#FAF2B3": 107,
    "#CFB88B": 108,
    "#898356": 110,
    "#744B1F": 111,
    "#C9D6A1": 112,
    "#95AE53": 114,
    "#516B3C": 116,
    "#262626": 123,
    "#F45212": 125,
    "#FC7373": 127,
    "#FF5B5B": 128,
    "#42ADF8": 130,
    "#0202F0": 133,
    "#69399F": 134,
    "#A58265": 135,
    "#B05E42": 136,
    "#E49704": 138,
    "#995BCD": 139,
    "#FB99B7": 140,
    "#9DBE8F": 141,
    "#B7673B": 144,
    "#911426": 145,
    "#57858F": 155,
    "#20912B": 150,
    "#8FD1DF": 151,
    "#874B2B": 147,
    "#595959": 156,
    "#2B4116": 157,
    "#1E366A": 159,
    "#7DE1B1": 161,
    "#3F0F04": 169,
    "#0B501B": 170,
    "#F20000": 233
}


# ========================
# 工具函数
# ========================
@st.cache_data
def init_preset_colors():
    """预处理颜色数据并缓存"""
    preset_colors = []
    for hex_color, num in COLOR_DICT.items():
        r = int(hex_color[1:3], 16)
        g = int(hex_color[3:5], 16)
        b = int(hex_color[5:7], 16)
        preset_colors.append(((r, g, b), num, hex_color))
    return preset_colors


@lru_cache(maxsize=1024)
def find_closest_color(pixel_rgb, _preset_colors):
    """带缓存的颜色匹配函数"""
    min_distance = float('inf')
    closest_num = 87  # 默认白色
    for preset in _preset_colors:
        r_diff = pixel_rgb[0] - preset[0][0]
        g_diff = pixel_rgb[1] - preset[0][1]
        b_diff = pixel_rgb[2] - preset[0][2]
        distance = r_diff ** 2 + g_diff ** 2 + b_diff ** 2
        if distance < min_distance:
            min_distance = distance
            closest_num = preset[1]
    return closest_num


# ========================
# 核心处理函数
# ========================
def process_image(image, option, max_blocks=None, custom_size=None):
    """处理图片核心逻辑"""
    try:
        img = image.convert("RGBA")
        original_width, original_height = img.size

        # 计算目标尺寸
        if option == "max_blocks":
            aspect_ratio = original_width / original_height
            new_width = int(math.sqrt(max_blocks * aspect_ratio))
            new_height = int(max_blocks / new_width)
        elif option == "custom_size":
            new_width, new_height = custom_size
        else:
            new_width, new_height = original_width, original_height

        img = img.resize((new_width, new_height))
        pixels = img.load()

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active

        # 设置单元格样式
        for col in range(1, new_width + 1):
            ws.column_dimensions[get_column_letter(col)].width = 2.86
        for row in range(1, new_height + 1):
            ws.row_dimensions[row].height = 15

        # 处理像素
        color_stats = {}
        preset_colors = init_preset_colors()

        for y in range(new_height):
            for x in range(new_width):
                r, g, b, a = pixels[x, y]

                # 处理透明像素
                if a < 255:
                    color_num = 87
                else:
                    color_num = find_closest_color((r, g, b), tuple(preset_colors))

                # 更新统计
                color_stats[color_num] = color_stats.get(color_num, 0) + 1

                # 写入单元格
                cell = ws.cell(row=y + 1, column=x + 1)
                cell.value = color_num
                if color_num != 87:
                    hex_color = next((c[2] for c in preset_colors if c[1] == color_num), "#FFFFFF")
                    cell.fill = PatternFill(
                        start_color=hex_color[1:],
                        end_color=hex_color[1:],
                        fill_type="solid"
                    )

        # 添加统计信息
        add_statistics(ws, new_width, color_stats, preset_colors)

        return wb, (new_width, new_height), color_stats

    except Exception as e:
        st.error(f"图片处理错误: {str(e)}")
        return None, None, None


def add_statistics(ws, width, color_stats, preset_colors):
    """添加统计信息和二维码到Excel"""
    start_col = width + 5

    # 统计表头
    ws.cell(row=1, column=start_col, value="积木编号")
    ws.cell(row=1, column=start_col + 1, value="数量")

    # 添加统计数据
    for row, (color_num, count) in enumerate(sorted(color_stats.items()), start=2):
        hex_color = next((c[2] for c in preset_colors if c[1] == color_num), "#FFFFFF")

        # 编号单元格
        num_cell = ws.cell(row=row, column=start_col)
        num_cell.value = color_num
        if color_num != 87:
            num_cell.fill = PatternFill(
                start_color=hex_color[1:],
                end_color=hex_color[1:],
                fill_type="solid"
            )

        # 数量单元格
        ws.cell(row=row, column=start_col + 1, value=count)

        # 修改后的二维码插入代码
try:
    from openpyxl.drawing.image import Image as xlImage

    # 构建GitHub图片URL（假设图片在main分支的根目录）
    github_url = "https://github.com/tudoudaren/jiafei/blob/main/taobao_qr.jpg"
    
    # 下载图片到临时文件
    with NamedTemporaryFile(delete=False, suffix=".jpg") as tmp_file:
        response = requests.get(github_url, timeout=10)
        response.raise_for_status()  # 检查HTTP错误
        tmp_file.write(response.content)
        tmp_path = tmp_file.name

    # 加载图片并调整尺寸
    qr_img = xlImage(tmp_path)
    qr_img.width = 150
    qr_img.height = 150

    # 定位到统计表右侧
    anchor_col = get_column_letter(start_col + 3)
    
    # 添加文字到第一行
    ws[f"{anchor_col}1"] = "淘宝店铺扫码"
    ws[f"{anchor_col}1"].font = Font(bold=True, color="FF4500")
    
    # 插入图片到第二行
    ws.add_image(qr_img, f"{anchor_col}2")
    
    # 设置行高确保显示
    ws.row_dimensions[2].height = 120  # 调整第二行高度

except requests.exceptions.RequestException as e:
    st.error(f"二维码下载失败: {str(e)}")
    print(f"网络请求异常: {str(e)}")
except Exception as e:
    st.error(f"二维码插入失败: {str(e)}")
    print(f"二维码处理异常: {str(e)}")
finally:
    # 清理临时文件
    if 'tmp_path' in locals() and os.path.exists(tmp_path):
        os.remove(tmp_path)
# ========================
# Streamlit界面
# ========================
def main():
    st.set_page_config(
        page_title="加飞积木-图片转Excel像素画",
        page_icon="🧱",
        layout="centered"
    )

    # 标题与二维码分栏布局
    col1, col2 = st.columns([2, 1])
    with col1:
        st.title("🧱图片转excel像素画-“加飞积木”制作")
        st.markdown("淘宝小红书关注“加飞积木”，B站关注“某山楂”")
    with col2:
        qr_image = Image.open("taobao_qr.jpg")  # 确保图片文件在项目目录下
        st.image(qr_image,
                caption="扫码淘宝店铺，或小红书关注“加飞积木”，可定制或批量采购",
                width=200,
                use_container_width=False)  # 已更新参数

    st.markdown("---")

    # 文件上传
    uploaded_file = st.file_uploader("1、点击Browse files上传图片", type=["jpg", "jpeg", "png"])
    if not uploaded_file:
        return

    # 处理选项
    option = st.radio("2、选择像素画尺寸", [
        "最大积木数量",
        "自定义尺寸",
        "原始尺寸"
    ], index=0)

    # 参数输入
    max_blocks = None
    custom_size = None

    if option == "最大积木数量":
        max_blocks = st.number_input("输入使用的最大积木数量", min_value=1, value=1000, step=100)
    elif option == "自定义尺寸":
        col1, col2 = st.columns(2)
        with col1:
            width = st.number_input("宽度（积木数）", min_value=1, value=100)
        with col2:
            height = st.number_input("高度（积木数）", min_value=1, value=100)
        custom_size = (width, height)

    # 处理按钮
    if st.button("一键生成像素画", use_container_width=True):
        with st.spinner("正在生成Excel文件..."):
            try:
                image = Image.open(uploaded_file)
                wb, size, stats = process_image(
                    image,
                    option="max_blocks" if option == "最大积木数量" else \
                        "custom_size" if option == "自定义尺寸" else "original",
                    max_blocks=max_blocks,
                    custom_size=custom_size
                )

                if wb:
                    st.success(f"转换成功！最终尺寸：{size[0]}x{size[1]} 像素")

                    # 生成下载文件
                    excel_buffer = io.BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)

                    # 生成带时间戳的文件名（新增部分）
                    from datetime import datetime
                    timestamp = datetime.now().strftime("%Y%m%d%H%M")
                    filename = f"加飞积木_像素画_{timestamp}.xlsx"

                    st.download_button(
                        label="下载Excel文件",
                        data=excel_buffer,
                        file_name=filename,  # 修改文件名
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

            except Exception as e:
                st.error(f"处理失败: {str(e)}")


if __name__ == "__main__":
    main()
